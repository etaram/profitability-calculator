import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import numpy_financial as npf
from scipy import optimize
from datetime import datetime
from dateutil.relativedelta import relativedelta

# ניסיון לייבא plotly
try:
    import plotly.graph_objects as go

    plotly_available = True
except ImportError:
    plotly_available = False
    st.warning("הספרייה 'plotly' לא מותקנת. חלק מהתרשימים לא יוצגו. להתקנה, הרץ 'pip install plotly'.")

# עיצוב כותרת האפליקציה והסגנון הכללי
st.markdown(
    """
    <style>
    .main-title {
        font-size: 42px;
        color: #d04f30;
        text-align: center;
        margin-bottom: 30px;
    }
    .stApp {
        background-color: #ffffff;
    }
    .sidebar .sidebar-content {
        background-color: #f0f0f0;
    }
    .stButton>button {
        background-color: #4CAF50;
        color: white;
        border-radius: 10px;
        border: none;
        height: 40px;
        width: 100%;
        margin-top: 10px;
    }
    .stButton>button:hover {
        background-color: #45a049;
        color: white;
    }
    .css-1cpxqw2, .css-1kyxreq, .css-14xtw13, .css-1lcbmhc, .css-2trqyj, .css-1vbd788, .css-1b0ba9k, .css-10trblm {
        color: #000000 !important;
        font-size: 20px !important;
        text-align: right;
    }
    .dataframe {
        background-color: #ffffff;
        color: #000000 !important;
        font-size: 20px !important;
        border: 1px solid #dddddd;
        text-align: right;
        padding: 5px;
        border-radius: 5px;
    }
    @media (max-width: 768px) {
        .stApp {
            padding: 20px;
        }
    }
    </style>
    """,
    unsafe_allow_html=True
)

# כותרת ראשית
st.markdown('<p class="main-title">מחשבון השקעה דינאמי לפרויקט וילות</p>', unsafe_allow_html=True)
st.image("קבוצת מעיינות.png", use_column_width=True)


def xnpv(rate, values, dates):
    if rate <= -1.0:
        return float('inf')
    d0 = dates[0]
    return sum(vi / (1.0 + rate) ** ((di - d0).days / 365.0) for vi, di in zip(values, dates))


def xirr(values, dates):
    try:
        return optimize.newton(lambda r: xnpv(r, values, dates), 0.0)
    except RuntimeError:
        return np.nan


# פונקציה לחישוב החזר חודשי לפי סוג סילוקין
def calculate_loan_repayment(total_loan, annual_rate, loan_term_years, repayment_type):
    monthly_rate = annual_rate / 12
    num_payments = loan_term_years * 12
    payments = []

    if repayment_type == 'שפיצר':
        monthly_payment = npf.pmt(monthly_rate, num_payments, -total_loan)
        for n in range(1, num_payments + 1):
            interest_payment = total_loan * monthly_rate
            principal_payment = monthly_payment - interest_payment
            total_loan -= principal_payment
            payments.append({
                'תשלום': n,
                'קרן לתשלום': round(principal_payment),
                'ריבית לתשלום': round(interest_payment),
                'תשלום חודשי': round(monthly_payment),
                'יתרת הלוואה': round(total_loan)
            })

    elif repayment_type == 'קרן שווה':
        monthly_principal = total_loan / num_payments
        for n in range(1, num_payments + 1):
            interest_payment = total_loan * monthly_rate
            total_payment = monthly_principal + interest_payment
            total_loan -= monthly_principal
            payments.append({
                'תשלום': n,
                'קרן לתשלום': round(monthly_principal),
                'ריבית לתשלום': round(interest_payment),
                'תשלום חודשי': round(total_payment),
                'יתרת הלוואה': round(total_loan)
            })

    elif repayment_type == 'בוליט':
        for n in range(1, num_payments + 1):
            interest_payment = total_loan * monthly_rate
            payments.append({
                'תשלום': n,
                'קרן לתשלום': 0 if n < num_payments else total_loan,
                'ריבית לתשלום': round(interest_payment),
                'תשלום חודשי': round(interest_payment) if n < num_payments else round(interest_payment + total_loan),
                'יתרת הלוואה': total_loan if n < num_payments else 0
            })

    return payments


# פונקציה לחישוב תוצאות פיננסיות לפרויקט
def calculate_financial_metrics(villas, size):
    # חישוב עלויות הקמה
    construction_cost = construction_cost_per_sqm * size * villas
    land_cost = land_cost_per_villa * villas
    total_construction_cost = (
            construction_cost +
            land_cost +
            land_development_cost +
            public_area_development_cost +
            reception_and_logistics_cost +
            small_event_hall_cost +
            planning_and_consultants_cost
    )

    # חישוב הכנסות שנתיות מהשכרת וילות
    annual_revenue = price_per_night * occupancy_rate * 365 * villas

    # חישוב רווח גולמי שנתי (ברוטו)
    gross_annual_profit = annual_revenue

    # חישוב עלויות תפעול משתנות (ניקיון ואביזרים)
    total_cleaning_cost = cleaning_cost_per_night * 365 * occupancy_rate * villas
    total_accessories_cost = accessories_cost_per_night * 365 * occupancy_rate * villas

    # חישוב עלויות תפעול קבועות (תפעול חודשי לוילה, ביטוח ושיווק)
    operational_cost_per_villa = monthly_operational_cost_per_villa * 12
    total_insurance_cost = annual_insurance_cost_per_villa * villas
    total_operational_fixed_cost = (operational_cost_per_villa * villas + total_insurance_cost + annual_marketing_cost)

    # חישוב רווח תפעולי שנתי
    operating_annual_profit = gross_annual_profit - (
                total_cleaning_cost + total_accessories_cost + total_operational_fixed_cost)

    # חישוב עלויות מימון
    total_loan = total_construction_cost - equity_amount
    loan_payments = calculate_loan_repayment(total_loan, prime_interest_rate + additional_interest_rate, loan_term,
                                             repayment_type)
    annual_financing_cost = sum(payment['ריבית לתשלום'] for payment in loan_payments[:12])  # סכום הריבית לשנה הראשונה

    # חישוב רווח נקי שנתי לפני סובסידיה (כולל מס שנתי ועלויות מימון)
    net_annual_profit_before_subsidy = (operating_annual_profit - annual_financing_cost) * (1 - tax_rate)

    # חישוב חלק שנתי של הסובסידיה
    annual_subsidy_min = (total_construction_cost * subsidy_min) / loan_term
    annual_subsidy_max = (total_construction_cost * subsidy_max) / loan_term

    # חישוב רווח נקי שנתי כולל סובסידיה (מחולק לאורך כל תקופת ההלוואה)
    net_annual_profit_with_subsidy_min = net_annual_profit_before_subsidy + annual_subsidy_min
    net_annual_profit_with_subsidy_max = net_annual_profit_before_subsidy + annual_subsidy_max

    # חישוב NPV עם סובסידיה וערך שארית
    terminal_value = net_annual_profit_with_subsidy_min * 10  # הנחה: ערך שארית הוא 10 שנות רווח
    cash_flow_min = [-total_construction_cost] + [net_annual_profit_with_subsidy_min] * loan_term + [terminal_value]
    cash_flow_max = [-total_construction_cost] + [net_annual_profit_with_subsidy_max] * loan_term + [terminal_value]

    dates = [datetime.now() + relativedelta(years=i) for i in
             range(loan_term + 2)]  # +2 for initial investment and terminal value

    total_npv_min = npf.npv(discount_rate, cash_flow_min)
    total_npv_max = npf.npv(discount_rate, cash_flow_max)

    # חישוב ROI משופר (XIRR)
    values_min = [-total_construction_cost * (1 - subsidy_min)] + [net_annual_profit_with_subsidy_min] * loan_term + [
        terminal_value]
    values_max = [-total_construction_cost * (1 - subsidy_max)] + [net_annual_profit_with_subsidy_max] * loan_term + [
        terminal_value]

    roi_min = xirr(values_min, dates) * 100
    roi_max = xirr(values_max, dates) * 100

    # חישוב IRR עם טיפול במקרי קצה
    irr_min = npf.irr(cash_flow_min)
    irr_max = npf.irr(cash_flow_max)

    irr_min = irr_min * 100 if not (np.isnan(irr_min) or np.isinf(irr_min)) else 0
    irr_max = irr_max * 100 if not (np.isnan(irr_max) or np.isinf(irr_max)) else 0

    # חישוב תקופת החזר
    payback_period_min = (total_construction_cost * (1 - subsidy_min)) / net_annual_profit_with_subsidy_min
    payback_period_max = (total_construction_cost * (1 - subsidy_max)) / net_annual_profit_with_subsidy_max

    # החזרת ערכי המדדים הפיננסיים
    return {
        'Gross Annual Profit': gross_annual_profit,
        'Operating Annual Profit': operating_annual_profit,
        'Net Annual Profit (Before Subsidy)': net_annual_profit_before_subsidy,
        'Net Annual Profit with Min Subsidy': net_annual_profit_with_subsidy_min,
        'Net Annual Profit with Max Subsidy': net_annual_profit_with_subsidy_max,
        'NPV Min': total_npv_min,
        'NPV Max': total_npv_max,
        'ROI Min': roi_min,
        'ROI Max': roi_max,
        'IRR Min': irr_min,
        'IRR Max': irr_max,
        'Payback Period Min': payback_period_min,
        'Payback Period Max': payback_period_max,
        'Total Construction Cost': total_construction_cost,
        'Annual Operational Cost': total_operational_fixed_cost + total_cleaning_cost + total_accessories_cost,
        'Annual Financing Cost': annual_financing_cost
    }


# קלטים מהמשתמש
num_villas = st.slider('מספר וילות', min_value=5, max_value=40, value=10, step=1)
villa_size_sqm = st.slider('גודל וילה (מ"ר)', min_value=120, max_value=250, value=200, step=10)
price_per_night = st.number_input('מחיר ללילה (₪)', min_value=1, value=3500)
occupancy_rate = st.slider('שיעור תפוסה (%)', min_value=0, max_value=100, value=40) / 100
land_cost_per_villa = st.number_input('עלות קרקע לוילה (₪)', min_value=0, value=500000)
construction_cost_per_sqm = st.number_input('עלות בנייה למ"ר (₪)', min_value=0, value=15000)
monthly_operational_cost_per_villa = st.number_input('עלות תפעול חודשית לוילה (₪)', min_value=0, value=4000)
annual_marketing_cost = st.number_input('עלות שיווק שנתית (₪)', min_value=0, value=600000)
land_development_cost = st.number_input('עלות פיתוח קרקע (₪)', min_value=0, value=200000)
public_area_development_cost = st.number_input('עלות פיתוח שטח ציבורי וחניות (₪)', min_value=0, value=1000000)
reception_and_logistics_cost = st.number_input('עלות מבנה קבלה ולוגיסטיקה (₪)', min_value=0, value=700000)
small_event_hall_cost = st.number_input('עלות אולם אירועים קטן (₪)', min_value=0, value=700000)
planning_and_consultants_cost = st.number_input('עלות תכנון ויועצים (₪)', min_value=0, value=150000)
cleaning_cost_per_night = st.number_input('עלות ניקיון ללילה (₪)', min_value=0, value=200)
accessories_cost_per_night = st.number_input('עלות אביזרים ללילה (₪)', min_value=0, value=50)
annual_insurance_cost_per_villa = st.number_input('עלות ביטוח שנתית לוילה (₪)', min_value=0, value=5000)
annual_inflation_rate = st.slider('שיעור אינפלציה שנתי (%)', min_value=0, max_value=100, value=2) / 100
discount_rate = st.slider('שיעור היוון (%)', min_value=0, max_value=100, value=8) / 100
prime_interest_rate = st.slider('ריבית פריים (%)', min_value=0.0, max_value=100.0, value=3.5) / 100
additional_interest_rate = st.number_input('ריבית נוספת מעל הפריים (%)', min_value=0.0, value=0.0) / 100
equity_amount = st.number_input('הון עצמי (₪)', min_value=0, value=1000000)
loan_term = st.slider('תקופת הלוואה (שנים)', min_value=5, max_value=15, value=15, step=1)
repayment_type = st.selectbox('סוג סילוקין', ['שפיצר', 'קרן שווה', 'בוליט'])
tax_rate = st.slider('שיעור מס (%)', min_value=0.0, max_value=50.0, value=7.5) / 100

# שיעור מס וסובסידיות
subsidy_min = 20 / 100
subsidy_max = 30 / 100

# חישוב עלות ההשקעה בניכוי הון עצמי
total_project_cost = (
        construction_cost_per_sqm * villa_size_sqm * num_villas +
        land_cost_per_villa * num_villas +
        land_development_cost +
        public_area_development_cost +
        reception_and_logistics_cost +
        small_event_hall_cost +
        planning_and_consultants_cost
)

loan_amount = total_project_cost - equity_amount

# חישוב תשלומי הלוואה
loan_payments = calculate_loan_repayment(loan_amount, prime_interest_rate + additional_interest_rate, loan_term,
                                         repayment_type)
loan_payments_df = pd.DataFrame(loan_payments)

# חישוב תוצאות פיננסיות לפרויקט
metrics = calculate_financial_metrics(num_villas, villa_size_sqm)


# פונקציה להצגת מדד צבעוני
def color_metric(value, threshold_good, threshold_bad, reverse=False):
    if reverse:
        if value <= threshold_good:
            return "green"
        elif value >= threshold_bad:
            return "red"
        else:
            return "orange"
    else:
        if value >= threshold_good:
            return "green"
        elif value <= threshold_bad:
            return "red"
        else:
            return "orange"


# הצגת מדדים צבעוניים
st.markdown("### מדדים עיקריים")

col1, col2, col3, col4 = st.columns(4)

with col1:
    roi_color = color_metric(metrics['ROI Min'], 15, 5)
    st.markdown(f"<h1 style='text-align: center; color: {roi_color};'>{metrics['ROI Min']:.2f}%</h1>",
                unsafe_allow_html=True)
    st.markdown("<p style='text-align: center;'>ROI מינימלי</p>", unsafe_allow_html=True)

with col2:
    npv_color = color_metric(metrics['NPV Min'], 0, -1000000)
    st.markdown(f"<h1 style='text-align: center; color: {npv_color};'>{int(metrics['NPV Min']):,} ₪</h1>",
                unsafe_allow_html=True)
    st.markdown("<p style='text-align: center;'>NPV מינימלי</p>", unsafe_allow_html=True)

with col3:
    irr_color = color_metric(metrics['IRR Min'], 10, 5)
    st.markdown(f"<h1 style='text-align: center; color: {irr_color};'>{metrics['IRR Min']:.2f}%</h1>",
                unsafe_allow_html=True)
    st.markdown("<p style='text-align: center;'>IRR מינימלי</p>", unsafe_allow_html=True)

with col4:
    payback_color = color_metric(metrics['Payback Period Max'], 7, 10, reverse=True)
    st.markdown(f"<h1 style='text-align: center; color: {payback_color};'>{metrics['Payback Period Max']:.2f}</h1>",
                unsafe_allow_html=True)
    st.markdown("<p style='text-align: center;'>תקופת החזר מקסימלית (שנים)</p>", unsafe_allow_html=True)

# הוספת הסבר על המדדים
st.markdown("""
**הסבר על המדדים:**
- **ROI (Return on Investment)**: אחוז התשואה על ההשקעה. ערך גבוה יותר טוב יותר.
- **NPV (Net Present Value)**: הערך הנוכחי הנקי של ההשקעה. ערך חיובי מצביע על השקעה כדאית.
- **IRR (Internal Rate of Return)**: שיעור התשואה הפנימי. ככל שהערך גבוה יותר, כך ההשקעה אטרקטיבית יותר.
- **תקופת החזר**: הזמן שלוקח להחזיר את ההשקעה הראשונית. ערך נמוך יותר טוב יותר.

שים לב: יש לבחון את כל המדדים יחד כדי לקבל תמונה מלאה על כדאיות ההשקעה.
""")

# הצגת תוצאות פיננסיות מפורטות
st.markdown("### תוצאות פיננסיות מפורטות")

col1, col2 = st.columns(2)

with col1:
    st.write(f"**רווח גולמי שנתי:** {int(metrics['Gross Annual Profit']):,} ₪")
    st.write(f"**רווח תפעולי שנתי:** {int(metrics['Operating Annual Profit']):,} ₪")
    st.write(f"**רווח נקי שנתי (לפני סובסידיה):** {int(metrics['Net Annual Profit (Before Subsidy)']):,} ₪")
    st.write(f"**רווח נקי שנתי (עם סובסידיה מינימלית):** {int(metrics['Net Annual Profit with Min Subsidy']):,} ₪")
    st.write(f"**רווח נקי שנתי (עם סובסידיה מקסימלית):** {int(metrics['Net Annual Profit with Max Subsidy']):,} ₪")

with col2:
    st.write(f"**ROI מינימלי:** {metrics['ROI Min']:.2f}%")
    st.write(f"**ROI מקסימלי:** {metrics['ROI Max']:.2f}%")
    st.write(f"**IRR מינימלי:** {metrics['IRR Min']:.2f}%")
    st.write(f"**IRR מקסימלי:** {metrics['IRR Max']:.2f}%")
    st.write(f"**NPV מינימלי:** {int(metrics['NPV Min']):,} ₪")
    st.write(f"**NPV מקסימלי:** {int(metrics['NPV Max']):,} ₪")

st.write(f"**תקופת החזר מינימלית:** {metrics['Payback Period Min']:.2f} שנים")
st.write(f"**תקופת החזר מקסימלית:** {metrics['Payback Period Max']:.2f} שנים")
st.write(f"**עלות הקמה כוללת:** {int(metrics['Total Construction Cost']):,} ₪")
st.write(f"**עלות תפעול שנתית:** {int(metrics['Annual Operational Cost']):,} ₪")
st.write(f"**עלות מימון שנתית:** {int(metrics['Annual Financing Cost']):,} ₪")

if plotly_available:
    # תרשים עוגה להשוואת חלוקת ההכנסות
    operating_profit = metrics['Operating Annual Profit']
    net_profit = metrics['Net Annual Profit (Before Subsidy)']
    financing_costs = metrics['Annual Financing Cost']

    fig_pie = go.Figure(data=[go.Pie(
        labels=['רווח נקי', 'עלויות תפעול', 'עלויות מימון', 'מיסים'],
        values=[net_profit,
                metrics['Annual Operational Cost'],
                financing_costs,
                (operating_profit - financing_costs) * tax_rate],
        hole=.3
    )])
    fig_pie.update_layout(title_text="התפלגות ההכנסות השנתיות")
    st.plotly_chart(fig_pie)

    # תרשים מפל להצגת השפעת הפרמטרים על הרווח
    fig_waterfall = go.Figure(go.Waterfall(
        name="20", orientation="v",
        measure=["relative", "relative", "relative", "relative", "total"],
        x=["הכנסות ברוטו", "עלויות תפעול", "עלויות מימון", "מיסים", "רווח נקי"],
        textposition="outside",
        text=[f"+{metrics['Gross Annual Profit']:,.0f}",
              f"-{metrics['Annual Operational Cost']:,.0f}",
              f"-{financing_costs:,.0f}",
              f"-{((operating_profit - financing_costs) * tax_rate):,.0f}",
              f"{net_profit:,.0f}"],
        y=[metrics['Gross Annual Profit'],
           -metrics['Annual Operational Cost'],
           -financing_costs,
           -((operating_profit - financing_costs) * tax_rate),
           net_profit],
        connector={"line": {"color": "rgb(63, 63, 63)"}},
    ))
    fig_waterfall.update_layout(title_text="ניתוח רווח שנתי", waterfallgap=0.3)
    st.plotly_chart(fig_waterfall)
else:
    st.warning("הספרייה 'plotly' לא מותקנת. התרשימים לא יוצגו.")

# הצגת טבלת תשלומי הלוואה
st.markdown("<div dir='rtl'>### תשלומי הלוואה לפי סוג סילוקין</div>", unsafe_allow_html=True)
st.dataframe(loan_payments_df.style.set_properties(**{'text-align': 'right'}))

# ניתוח רגישות
st.markdown("### ניתוח רגישות")

# בחירת משתנה לניתוח
variable_to_analyze = st.selectbox('בחר משתנה לניתוח רגישות',
                                   ['מספר וילות', 'שיעור תפוסה', 'מחיר ללילה', 'שיעור היוון', 'שיעור מס'])

# הגדרת טווחים
if variable_to_analyze == 'מספר וילות':
    variable_range = range(5, 41)
elif variable_to_analyze == 'שיעור תפוסה':
    variable_range = range(20, 101)
elif variable_to_analyze == 'מחיר ללילה':
    variable_range = range(2000, 8001, 100)
elif variable_to_analyze == 'שיעור היוון':
    variable_range = range(1, 21)
elif variable_to_analyze == 'שיעור מס':
    variable_range = range(0, 51)


# פונקציה לניתוח רגישות
def sensitivity_analysis(base_villas, base_size, variable_name, variable_range):
    results = []
    for value in variable_range:
        if variable_name == 'מספר וילות':
            temp_metrics = calculate_financial_metrics(value, base_size)
        elif variable_name == 'שיעור תפוסה':
            global occupancy_rate
            occupancy_rate = value / 100
            temp_metrics = calculate_financial_metrics(base_villas, base_size)
        elif variable_name == 'מחיר ללילה':
            global price_per_night
            price_per_night = value
            temp_metrics = calculate_financial_metrics(base_villas, base_size)
        elif variable_name == 'שיעור היוון':
            global discount_rate
            discount_rate = value / 100
            temp_metrics = calculate_financial_metrics(base_villas, base_size)
        elif variable_name == 'שיעור מס':
            global tax_rate
            tax_rate = value / 100
            temp_metrics = calculate_financial_metrics(base_villas, base_size)

        results.append({
            'ערך': value,
            'ROI': temp_metrics['ROI Min'],
            'NPV': temp_metrics['NPV Min'],
            'IRR': temp_metrics['IRR Min'],
            'תקופת החזר': temp_metrics['Payback Period Max']
        })

    return results

# ביצוע ניתוח רגישות
sensitivity_results = sensitivity_analysis(num_villas, villa_size_sqm, variable_to_analyze, variable_range)
df_sensitivity = pd.DataFrame(sensitivity_results)

# הצגת גרף ניתוח רגישות
if plotly_available:
    fig_sensitivity = go.Figure()

    fig_sensitivity.add_trace(go.Scatter(x=df_sensitivity['ערך'], y=df_sensitivity['ROI'],
                                         mode='lines+markers', name='ROI (%)'))
    fig_sensitivity.add_trace(go.Scatter(x=df_sensitivity['ערך'], y=df_sensitivity['NPV'] / 1_000_000,
                                         mode='lines+markers', name='NPV (מיליוני ₪)'))
    fig_sensitivity.add_trace(go.Scatter(x=df_sensitivity['ערך'], y=df_sensitivity['IRR'],
                                         mode='lines+markers', name='IRR (%)'))
    fig_sensitivity.add_trace(go.Scatter(x=df_sensitivity['ערך'], y=df_sensitivity['תקופת החזר'],
                                         mode='lines+markers', name='תקופת החזר (שנים)'))

    fig_sensitivity.update_layout(title=f'ניתוח רגישות: השפעת {variable_to_analyze} על מדדים פיננסיים',
                                  xaxis_title=variable_to_analyze,
                                  yaxis_title='ערך')

    st.plotly_chart(fig_sensitivity)
else:
    st.warning("הספרייה 'plotly' לא מותקנת. גרף ניתוח הרגישות לא יוצג.")

# הצגת הנתונים בטבלה
st.dataframe(df_sensitivity)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter


def generate_advanced_excel_report():
    wb = openpyxl.Workbook()

    # גיליון נתונים
    ws_data = wb.active
    ws_data.title = "נתונים"

    # כותרות
    headers = [
        "פרמטר", "ערך", "יחידה",
        "מספר וילות", "גודל וילה", "מחיר ללילה", "שיעור תפוסה",
        "עלות קרקע לוילה", "עלות בנייה למ\"ר", "עלות תפעול חודשית לוילה",
        "עלות שיווק שנתית", "עלות פיתוח קרקע", "עלות פיתוח שטח ציבורי",
        "עלות מבנה קבלה", "עלות אולם אירועים", "עלות תכנון ויועצים",
        "עלות ניקיון ללילה", "עלות אביזרים ללילה", "עלות ביטוח שנתית לוילה",
        "שיעור אינפלציה שנתי", "שיעור היוון", "ריבית פריים",
        "ריבית נוספת מעל הפריים", "הון עצמי", "תקופת הלוואה", "סוג סילוקין",
        "שיעור מס"
    ]

    for col, header in enumerate(headers, start=1):
        ws_data.cell(row=1, column=col, value=header)
        ws_data.cell(row=1, column=col).font = Font(bold=True)

    # נתונים
    data = [
        ("מספר וילות", num_villas, ""),
        ("גודל וילה", villa_size_sqm, "מ\"ר"),
        ("מחיר ללילה", price_per_night, "₪"),
        ("שיעור תפוסה", occupancy_rate * 100, "%"),
        ("עלות קרקע לוילה", land_cost_per_villa, "₪"),
        ("עלות בנייה למ\"ר", construction_cost_per_sqm, "₪"),
        ("עלות תפעול חודשית לוילה", monthly_operational_cost_per_villa, "₪"),
        ("עלות שיווק שנתית", annual_marketing_cost, "₪"),
        ("עלות פיתוח קרקע", land_development_cost, "₪"),
        ("עלות פיתוח שטח ציבורי", public_area_development_cost, "₪"),
        ("עלות מבנה קבלה", reception_and_logistics_cost, "₪"),
        ("עלות אולם אירועים", small_event_hall_cost, "₪"),
        ("עלות תכנון ויועצים", planning_and_consultants_cost, "₪"),
        ("עלות ניקיון ללילה", cleaning_cost_per_night, "₪"),
        ("עלות אביזרים ללילה", accessories_cost_per_night, "₪"),
        ("עלות ביטוח שנתית לוילה", annual_insurance_cost_per_villa, "₪"),
        ("שיעור אינפלציה שנתי", annual_inflation_rate * 100, "%"),
        ("שיעור היוון", discount_rate * 100, "%"),
        ("ריבית פריים", prime_interest_rate * 100, "%"),
        ("ריבית נוספת מעל הפריים", additional_interest_rate * 100, "%"),
        ("הון עצמי", equity_amount, "₪"),
        ("תקופת הלוואה", loan_term, "שנים"),
        ("סוג סילוקין", repayment_type, ""),
        ("שיעור מס", tax_rate * 100, "%")
    ]

    for row, (param, value, unit) in enumerate(data, start=2):
        ws_data.cell(row=row, column=1, value=param)
        ws_data.cell(row=row, column=2, value=value)
        ws_data.cell(row=row, column=3, value=unit)

    # גיליון חישובים
    ws_calc = wb.create_sheet(title="חישובים")

    calc_headers = [
        "מדד", "ערך", "יחידה"
    ]

    for col, header in enumerate(calc_headers, start=1):
        ws_calc.cell(row=1, column=col, value=header)
        ws_calc.cell(row=1, column=col).font = Font(bold=True)

    calc_data = [
        ("רווח גולמי שנתי", metrics['Gross Annual Profit'], "₪"),
        ("רווח תפעולי שנתי", metrics['Operating Annual Profit'], "₪"),
        ("רווח נקי שנתי (לפני סובסידיה)", metrics['Net Annual Profit (Before Subsidy)'], "₪"),
        ("רווח נקי שנתי (עם סובסידיה מינימלית)", metrics['Net Annual Profit with Min Subsidy'], "₪"),
        ("רווח נקי שנתי (עם סובסידיה מקסימלית)", metrics['Net Annual Profit with Max Subsidy'], "₪"),
        ("ROI מינימלי", metrics['ROI Min'], "%"),
        ("ROI מקסימלי", metrics['ROI Max'], "%"),
        ("IRR מינימלי", metrics['IRR Min'], "%"),
        ("IRR מקסימלי", metrics['IRR Max'], "%"),
        ("NPV מינימלי", metrics['NPV Min'], "₪"),
        ("NPV מקסימלי", metrics['NPV Max'], "₪"),
        ("תקופת החזר מינימלית", metrics['Payback Period Min'], "שנים"),
        ("תקופת החזר מקסימלית", metrics['Payback Period Max'], "שנים"),
        ("עלות הקמה כוללת", metrics['Total Construction Cost'], "₪"),
        ("עלות תפעול שנתית", metrics['Annual Operational Cost'], "₪"),
        ("עלות מימון שנתית", metrics['Annual Financing Cost'], "₪")
    ]

    for row, (metric, value, unit) in enumerate(calc_data, start=2):
        ws_calc.cell(row=row, column=1, value=metric)
        ws_calc.cell(row=row, column=2, value=value)
        ws_calc.cell(row=row, column=3, value=unit)

    # גיליון גרף
    ws_chart = wb.create_sheet(title="גרף כדאיות")

    # נתונים לגרף
    chart_data = [
                     ("שנה", "תזרים מזומנים", "NPV מצטבר")
                 ] + [
                     (year,
                      metrics['Net Annual Profit with Min Subsidy'],
                      metrics['NPV Min'] * (1 - (1 / (1 + discount_rate)) ** (year)) / (1 - (1 / (1 + discount_rate))))
                     for year in range(1, loan_term + 1)
                 ]

    for row, data in enumerate(chart_data, start=1):
        for col, value in enumerate(data, start=1):
            ws_chart.cell(row=row, column=col, value=value)

    # יצירת גרף
    chart = LineChart()
    chart.title = "כדאיות הפרויקט לאורך זמן"
    chart.x_axis.title = "שנה"
    chart.y_axis.title = "ערך (₪)"

    data = Reference(ws_chart, min_col=2, min_row=1, max_col=3, max_row=loan_term + 1)
    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=loan_term + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws_chart.add_chart(chart, "E5")

    # שמירת הקובץ
    excel_filename = "advanced_investment_report.xlsx"
    wb.save(excel_filename)

    # הצגת הודעה למשתמש והורדת הקובץ
    with open(excel_filename, "rb") as file:
        st.download_button(label='Download Advanced Excel Report', data=file, file_name=excel_filename)


# עדכון הכפתור ליצירת דוח Excel
if st.button('Generate Advanced Excel Report'):
    generate_advanced_excel_report()

# סיום הקוד
st.markdown("---")
st.markdown("© 2024 מחשבון השקעה דינאמי . כל הזכויות שמורות לקבוצת מעיינות.")
